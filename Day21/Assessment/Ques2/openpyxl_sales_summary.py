from openpyxl import load_workbook, Workbook

file_name = "sales_data.xlsx"
output_file = "sales_summary.xlsx"
wb = load_workbook(file_name)
sheet = wb["2025"]

new_wb = Workbook()
new_sheet = new_wb.active
new_sheet.title = "Sales Summary"

headers = ["Product", "Quantity", "Price", "Total"]
new_sheet.append(headers)

for row in sheet.iter_rows(min_row=2, values_only=True):
    product, quantity, price = row
    total = quantity * price
    new_sheet.append([product, quantity, price, total])

new_wb.save(output_file)

print(f"File saved successfully as '{output_file}' using OpenPyXL.")
